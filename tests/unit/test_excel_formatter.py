"""
Testes unitários para o Excel Formatter — TASK-42.
"""
import io
from datetime import datetime
from typing import Any
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from orders_master.constants import Columns, GroupLabels, Highlight
from orders_master.formatting.excel_formatter import (
    build_excel,
    compute_scope_tag,
    sanitize_filename,
)


def test_sanitize_filename() -> None:
    assert sanitize_filename("Laboratório Mylan!") == "LaboratrioMylan"
    assert sanitize_filename("Labs-3_2026") == "Labs-3_2026"


def test_compute_scope_tag() -> None:
    # 1. TXT ativo
    assert compute_scope_tag(["Lab A"], object(), 47) == "TXT-47"
    
    # 2. 1 Laboratório
    assert compute_scope_tag(["Mylan"], None, 0) == "Mylan"
    
    # 3. Múltiplos Laboratórios
    assert compute_scope_tag(["Lab A", "Lab B", "Lab C"], None, 0) == "Labs-3"
    
    # 4. Sem filtro (Grupo)
    assert compute_scope_tag([], None, 0) == "GRUPO"


@patch("orders_master.formatting.excel_formatter.datetime")
def test_build_excel_filename(mock_datetime: Any) -> None:
    mock_datetime.now.return_value = datetime(2026, 5, 4, 14, 30)
    df = pd.DataFrame({Columns.CODIGO: [1001]})
    
    _, filename = build_excel(df, "Mylan")
    assert filename == "Sell_Out_Mylan_20260504_1430.xlsx"


def test_build_excel_output_valid() -> None:
    df = pd.DataFrame({Columns.CODIGO: [1001], Columns.STOCK: [10]})
    content, _ = build_excel(df, "TEST")
    
    # Verificar se é um ficheiro Excel válido carregável pelo openpyxl
    wb = load_workbook(io.BytesIO(content))
    assert "Propostas" in wb.sheetnames
    ws = wb["Propostas"]
    assert ws.cell(row=2, column=1).value == 1001


def test_apply_excel_rules_formatting() -> None:
    """
    Verifica se as cores do rules.py são aplicadas corretamente no Excel.
    """
    df = pd.DataFrame([
        # 1. Linha Grupo -> Fundo Preto
        {
            Columns.CODIGO: 1,
            Columns.LOCALIZACAO: GroupLabels.GROUP_ROW,
            Columns.STOCK: 100,
            Columns.DTVAL: "12/2030",
            Columns.PRICE_ANOMALY: False,
        },
        # 2. Não Comprar (DATA_OBS) -> Fundo Roxo (E6D5F5)
        {
            Columns.CODIGO: 2001,
            Columns.LOCALIZACAO: "FAR_A",
            Columns.STOCK: 10,
            Columns.DATA_OBS: "2026-01-01",
            Columns.T_UNI: 5,
            Columns.DTVAL: "12/2030",
            Columns.PRICE_ANOMALY: False,
        },
        # 3. Validade Curta -> Fundo Laranja (FFA500)
        {
            Columns.CODIGO: 3001,
            Columns.LOCALIZACAO: "FAR_A",
            Columns.STOCK: 10,
            Columns.DTVAL: datetime.now().strftime("%m/%Y"), # Validade este mês
            Columns.PRICE_ANOMALY: False,
        }
    ])
    
    content, _ = build_excel(df, "TEST")
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    # 1. Grupo (Fundo Preto: pode ser 00000000 ou FF000000 conforme a versão)
    black_fill = ws.cell(row=2, column=1).fill.start_color.index
    assert str(black_fill).endswith("000000")
    
    # 2. Não Comprar (Roxo: E6D5F5 -> FFE6D5F5)
    purple_fill = ws.cell(row=3, column=1).fill.start_color.index
    assert str(purple_fill).endswith("E6D5F5")
    
    # 3. Validade Curta (Laranja: FFA500 -> FFFFA500)
    # Coluna DTVAL é a 4ª neste DF (CODIGO, LOCALIZACAO, STOCK, DTVAL)
    orange_fill = ws.cell(row=4, column=4).fill.start_color.index
    assert str(orange_fill).endswith("FFA500")


def test_excel_precedence_grupo_over_rules() -> None:
    """
    Verifica que na linha Grupo as outras regras não se sobrepõem à cor preta.
    """
    df = pd.DataFrame([{
        Columns.CODIGO: 1,
        Columns.LOCALIZACAO: GroupLabels.GROUP_ROW,
        Columns.DATA_OBS: "2026-01-01", # Ativaria Não Comprar
        Columns.STOCK: 100,
    }])
    
    content, _ = build_excel(df, "TEST")
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    # Deve continuar preto (endswith 000000), não roxo.
    assert str(ws.cell(row=2, column=1).fill.start_color.index).endswith("000000")
