from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill

from orders_master.constants import Columns, GroupLabels
from orders_master.formatting.rules import RULES


def test_rules_definitions():
    assert len(RULES) == 5

    precedences = [r.precedence for r in RULES]
    assert precedences == [1, 2, 3, 4, 5]

    rule_grupo = next(r for r in RULES if r.name == "Grupo")
    assert isinstance(rule_grupo.excel_fill, PatternFill)
    assert isinstance(rule_grupo.excel_font, Font)
    assert "background-color" in rule_grupo.css_web

    rule_nao_comprar = next(r for r in RULES if r.name == "Não Comprar")
    assert isinstance(rule_nao_comprar.excel_fill, PatternFill)


def test_predicates():
    rule_grupo = next(r for r in RULES if r.name == "Grupo")
    s_grupo = pd.Series({Columns.LOCALIZACAO: GroupLabels.GROUP_ROW})
    assert rule_grupo.predicate(s_grupo)
    s_normal = pd.Series({Columns.LOCALIZACAO: "Farmácia A"})
    assert not rule_grupo.predicate(s_normal)

    rule_nao_comprar = next(r for r in RULES if r.name == "Não Comprar")
    assert rule_nao_comprar.predicate(pd.Series({Columns.DATA_OBS: "01-01-2026"}))
    assert not rule_nao_comprar.predicate(pd.Series({Columns.DATA_OBS: pd.NA}))

    rule_rutura = next(r for r in RULES if r.name == "Rutura")
    assert rule_rutura.predicate(pd.Series({Columns.DIR: "01-01-2026"}))
    assert not rule_rutura.predicate(pd.Series({Columns.DIR: pd.NA}))

    rule_validade = next(r for r in RULES if r.name == "Validade Curta")
    today = datetime.now()
    exp_month = (today.month % 12) + 1
    exp_year = today.year + (today.month // 12)
    s_validade_curta = pd.Series({Columns.DTVAL: f"{exp_month:02d}/{exp_year}"})
    assert rule_validade.predicate(s_validade_curta)

    s_validade_longa = pd.Series({Columns.DTVAL: "12/2099"})
    assert not rule_validade.predicate(s_validade_longa)

    rule_preco = next(r for r in RULES if r.name == "Preço Anómalo")
    assert rule_preco.predicate(pd.Series({Columns.PRICE_ANOMALY: True}))
    assert not rule_preco.predicate(pd.Series({Columns.PRICE_ANOMALY: False}))


def test_target_cells():
    df = pd.DataFrame(
        columns=[
            Columns.CODIGO,
            "X",
            "Y",
            Columns.T_UNI,
            Columns.PROPOSTA,
            Columns.DTVAL,
            Columns.PVP_MEDIO,
        ]
    )

    rule_grupo = next(r for r in RULES if r.name == "Grupo")
    assert rule_grupo.target_cells(df) == list(df.columns)

    rule_nao_comprar = next(r for r in RULES if r.name == "Não Comprar")
    assert rule_nao_comprar.target_cells(df) == [Columns.CODIGO, "X", "Y", Columns.T_UNI]

    rule_rutura = next(r for r in RULES if r.name == "Rutura")
    assert rule_rutura.target_cells(df) == [Columns.PROPOSTA]

    rule_validade = next(r for r in RULES if r.name == "Validade Curta")
    assert rule_validade.target_cells(df) == [Columns.DTVAL]

    rule_preco = next(r for r in RULES if r.name == "Preço Anómalo")
    assert rule_preco.target_cells(df) == [Columns.PVP_MEDIO]
