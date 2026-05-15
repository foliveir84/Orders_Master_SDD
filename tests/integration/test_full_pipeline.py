"""
tests/integration/test_full_pipeline.py — TASK-38.

Teste de integração end-to-end que valida o pipeline completo:
Ingestão -> Agregação -> Recálculo -> Exportação Excel.
"""

import io
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from orders_master.app_services.recalc_service import recalculate_proposal
from orders_master.app_services.session_service import process_orders_session
from orders_master.app_services.session_state import SessionState
from orders_master.constants import Columns, GroupLabels
from orders_master.formatting.excel_formatter import build_excel


@pytest.fixture
def mock_infoprex_content_f1() -> bytes:
    """Gera conteúdo para a Farmácia A."""
    header = "CPR\tNOM\tLOCALIZACAO\tSAC\tPVP\tPCU\tDUC\tDTVAL\tCLA\tDUV\tV0\tV1\tV2\tV3\n"
    # Produto 1: 10 unidades em stock, vende 5/mês. CLA=L1
    row1 = "2000001\tPRODUTO 1\tFARMACIA_A\t10\t10.0\t8.0\t100\t12/2026\tL1\t15/05/2024\t5\t5\t5\t5\n"
    return (header + row1).encode("utf-16")


@pytest.fixture
def mock_infoprex_content_f2() -> bytes:
    """Gera conteúdo para a Farmácia B."""
    header = "CPR\tNOM\tLOCALIZACAO\tSAC\tPVP\tPCU\tDUC\tDTVAL\tCLA\tDUV\tV0\tV1\tV2\tV3\n"
    # Produto 1: 5 unidades em stock, vende 5/mês. CLA=L1
    row1 = "2000001\tPRODUTO 1\tFARMACIA_B\t5\t10.0\t8.0\t100\t12/2026\tL1\t15/05/2024\t5\t5\t5\t5\n"
    return (header + row1).encode("utf-16")


@pytest.fixture
def labs_config_mock() -> MagicMock:
    """Mock do labs_config para process_orders_session."""
    mock = MagicMock()
    mock.root = {"LAB1": ["L1"]}
    return mock


def test_full_pipeline_aggregated(
    mock_infoprex_content_f1, mock_infoprex_content_f2, labs_config_mock
) -> None:
    """Valida o pipeline completo em modo AGRUPADO."""
    state = SessionState()

    file1 = io.BytesIO(mock_infoprex_content_f1)
    file1.name = "f1.txt"
    file2 = io.BytesIO(mock_infoprex_content_f2)
    file2.name = "f2.txt"

    # 1. Ingestão (com patch de secrets para evitar requests reais)
    with patch("orders_master.secrets_loader.get_secret", return_value=None):
        process_orders_session(
            files=[file1, file2],
            codes_file=None,
            brands_files=[],
            labs_selected=["LAB1"],
            labs_config=labs_config_mock,
            locations_aliases={"FARMACIA_A": "Loja A", "FARMACIA_B": "Loja B"},
            state=state,
        )

    assert len(state.file_inventory) == 2
    assert state.df_raw[Columns.LOCALIZACAO].nunique() == 2

    # 2. Recálculo Agrupado
    # Média mensal por loja = 5. Total média = 10.
    # Stock total = 15.
    # Previsão 2 meses -> 20. Proposta deve ser 20 - 15 = 5.
    df_final = recalculate_proposal(
        df_detailed=state.df_raw,
        detailed_view=False,
        df_master_products=state.df_master_products,
        months=2.0,
        weights=(1.0, 0.0, 0.0, 0.0),
        use_previous_month=False,
    )

    assert len(df_final) == 1
    assert df_final.iloc[0][Columns.PROPOSTA] == 5
    assert df_final.iloc[0][Columns.STOCK] == 15

    # 3. Exportação Excel
    excel_bytes, _ = build_excel(df_final, "TEST")
    assert len(excel_bytes) > 0

    # 4. Verificar Excel via openpyxl
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active
    assert ws.title == "Propostas"

    # Verificar cabeçalhos no Excel
    headers = [cell.value for cell in ws[1]]
    assert Columns.PROPOSTA in headers
    assert Columns.CODIGO in headers
    # Colunas técnicas devem estar ausentes (TASK-38 critério)
    assert "_sort_key" not in headers
    assert "CÓDIGO_STR" not in headers


def test_full_pipeline_detailed(
    mock_infoprex_content_f1, mock_infoprex_content_f2, labs_config_mock
) -> None:
    """Valida o pipeline completo em modo DETALHADO."""
    state = SessionState()

    file1 = io.BytesIO(mock_infoprex_content_f1)
    file1.name = "f1.txt"
    file2 = io.BytesIO(mock_infoprex_content_f2)
    file2.name = "f2.txt"

    with patch("orders_master.secrets_loader.get_secret", return_value=None):
        process_orders_session(
            files=[file1, file2],
            codes_file=None,
            brands_files=[],
            labs_selected=["LAB1"],
            labs_config=labs_config_mock,
            locations_aliases={"FARMACIA_A": "Loja A", "FARMACIA_B": "Loja B"},
            state=state,
        )

    # 2. Recalc Detalhado
    df_final = recalculate_proposal(
        df_detailed=state.df_raw,
        detailed_view=True,
        df_master_products=state.df_master_products,
        months=2.0,
        weights=(1.0, 0.0, 0.0, 0.0),
        use_previous_month=False,
    )

    # Deve ter 3 linhas: Loja A, Loja B e Grupo
    assert len(df_final) == 3
    assert GroupLabels.GROUP_ROW in df_final[Columns.LOCALIZACAO].values

    # Verificar soma na linha Grupo
    grupo_row = df_final[df_final[Columns.LOCALIZACAO] == GroupLabels.GROUP_ROW].iloc[0]
    assert grupo_row[Columns.PROPOSTA] == 5
    assert grupo_row[Columns.STOCK] == 15

    # 3. Exportação Excel
    excel_bytes, _ = build_excel(df_final, "TEST_DET")
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    # Verificar se a linha Grupo está no Excel (linha 4, pois header + 2 lojas + 1 grupo?)
    # A ordem de agregação coloca o Grupo no fim.
    last_row_values = [cell.value for cell in ws[ws.max_row]]
    assert GroupLabels.GROUP_ROW in last_row_values
