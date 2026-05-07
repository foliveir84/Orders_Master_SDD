"""
tests/unit/test_web_excel_parity.py — TASK-43.

Guardião da regra SSOT: valida que `build_styler()` (web) e `build_excel()` (Excel)
produzem exactamente as mesmas cores de fundo nas mesmas células.

Estratégia:
1. Construir um DataFrame-fixture cobrindo todos os 5 cenários de regra.
2. Extrair cores do Styler via `Styler.export()` (dicts de estilo por célula).
3. Extrair cores do workbook openpyxl via `ws.cell(row, col).fill.fgColor.rgb`.
4. Normalizar ambas para hex RRGGBB maiúsculo e comparar.

Se uma nova regra for adicionada a `rules.py` sem actualizar ambos os renderers,
este teste falha com uma mensagem clara indicando qual célula diverge.
"""

from __future__ import annotations

import io
import re
from datetime import datetime, timedelta
from typing import Any

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from orders_master.constants import Columns, GroupLabels
from orders_master.formatting.excel_formatter import apply_excel_rules, build_excel
from orders_master.formatting.rules import RULES
from orders_master.formatting.web_styler import build_styler


# ---------------------------------------------------------------------------
# Helpers de normalização de cor
# ---------------------------------------------------------------------------


def _normalise_hex(raw: str) -> str | None:
    """
    Normaliza qualquer string de cor para hex RRGGBB maiúsculo sem '#'.

    Aceita:
    - '#RRGGBB', '#RGB', 'RRGGBB', 'FFRRGGBB' (openpyxl alpha-prefixed)
    - 'rgb(r, g, b)' do CSS
    Devolve None se não for reconhecível como cor sólida.
    """
    if not raw or raw.strip() in ("", "none", "transparent", "00000000"):
        return None

    raw = raw.strip()

    # openpyxl: alpha-prefixed FFRRGGBB
    if re.fullmatch(r"[0-9A-Fa-f]{8}", raw):
        alpha = raw[:2].upper()
        if alpha == "00":
            return None  # fully transparent
        return raw[2:].upper()

    # '#RRGGBB' or 'RRGGBB'
    m = re.fullmatch(r"#?([0-9A-Fa-f]{6})", raw)
    if m:
        return m.group(1).upper()

    # '#RGB'
    m = re.fullmatch(r"#([0-9A-Fa-f]{3})", raw)
    if m:
        c = m.group(1)
        return (c[0] * 2 + c[1] * 2 + c[2] * 2).upper()

    # CSS rgb(r, g, b)
    m = re.fullmatch(r"rgb\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)", raw, re.IGNORECASE)
    if m:
        return "{:02X}{:02X}{:02X}".format(int(m.group(1)), int(m.group(2)), int(m.group(3)))

    return None


def _css_bg_color(css: str) -> str | None:
    """Extrai o valor de background-color de uma string CSS de estilo."""
    m = re.search(r"background-color\s*:\s*([^;]+)", css, re.IGNORECASE)
    if m:
        return _normalise_hex(m.group(1).strip())
    return None


# ---------------------------------------------------------------------------
# Fixture DataFrame — cobre todos os 5 cenários
# ---------------------------------------------------------------------------

# Próxima data de validade daqui a ~2 meses (< 4 meses → aciona regra Validade Curta)
_future = datetime.now() + timedelta(days=45)
_dtval_short = f"{_future.month}/{_future.year}"


@pytest.fixture(scope="module")
def parity_df() -> pd.DataFrame:
    """
    DataFrame com uma linha por cenário de regra:
    0 — Grupo (row de agregação)
    1 — Não Comprar (DATA_OBS preenchida)
    2 — Rutura (DIR preenchida)
    3 — Validade Curta (DTVAL ≤ 4 meses)
    4 — Preço Anómalo (price_anomaly = True)
    5 — Normal (sem nenhuma regra activa)
    """
    data: dict[str, Any] = {
        str(Columns.CODIGO): [1, 2, 3, 4, 5, 6],
        str(Columns.DESIGNACAO): ["Grupo Prod", "Nao Comprar Prod", "Rutura Prod", "Validade Prod", "Anomalo Prod", "Normal Prod"],
        str(Columns.LOCALIZACAO): [GroupLabels.GROUP_ROW, "Farm A", "Farm B", "Farm C", "Farm D", "Farm E"],
        str(Columns.STOCK): [0, 5, 3, 8, 4, 10],
        str(Columns.T_UNI): [100, 50, 30, 60, 20, 80],
        str(Columns.PROPOSTA): [0, 5, 10, 2, 4, 8],
        str(Columns.MARCA): ["", "M1", "M2", "M3", "M4", "M5"],
        str(Columns.DATA_OBS): [pd.NA, "10-05-2024", pd.NA, pd.NA, pd.NA, pd.NA],
        str(Columns.DIR): [pd.NA, pd.NA, "2024-04-01", pd.NA, pd.NA, pd.NA],
        str(Columns.DPR): [pd.NA, pd.NA, "2024-07-01", pd.NA, pd.NA, pd.NA],
        str(Columns.TIME_DELTA): [pd.NA, pd.NA, 30, pd.NA, pd.NA, pd.NA],
        str(Columns.DTVAL): [pd.NA, pd.NA, pd.NA, _dtval_short, pd.NA, pd.NA],
        str(Columns.PVP_MEDIO): [0.0, 12.5, 8.0, 5.5, 999.0, 15.0],
        str(Columns.PRICE_ANOMALY): [False, False, False, False, True, False],
        str(Columns.SORT_KEY): [1, 0, 0, 0, 0, 0],
    }
    return pd.DataFrame(data)


# ---------------------------------------------------------------------------
# Extracção de cores do Styler (Web)
# ---------------------------------------------------------------------------


def _extract_web_bg_colors(df: pd.DataFrame) -> dict[tuple[int, str], str | None]:
    """
    Devolve dict {(row_idx, col_name): hex_bg | None} para cada célula.
    Usa Styler.export() que devolve os estilos aplicados (não requer render HTML).
    """
    styler = build_styler(df)
    # _compute() processa todos os apply()s e popula _styles
    styler._compute()  # type: ignore[attr-defined]

    result: dict[tuple[int, str], str | None] = {}
    col_names = list(df.columns)
    n_cols = len(col_names)

    ctx = styler.ctx  # type: ignore[attr-defined]

    for (row_i, col_j), styles in ctx.items():
        if col_j >= n_cols:
            continue
        col_name = col_names[col_j]
        # styles é lista de (attr, value) tuples
        bg = None
        for attr, val in styles:
            if attr == "background-color":
                bg = _normalise_hex(val)
                break
        result[(row_i, col_name)] = bg

    # Garantir que todas as células (sem estilo) também estão no dict
    for ri in range(len(df)):
        for cn in col_names:
            if (ri, cn) not in result:
                result[(ri, cn)] = None

    return result


# ---------------------------------------------------------------------------
# Extracção de cores do Excel (openpyxl)
# ---------------------------------------------------------------------------


def _extract_excel_bg_colors(df: pd.DataFrame) -> dict[tuple[int, str], str | None]:
    """
    Devolve dict {(row_idx, col_name): hex_bg | None} para cada célula.
    Usa build_excel() e depois abre com openpyxl.
    """
    excel_bytes, _ = build_excel(df, scope_tag="parity_test")
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    col_names = list(df.columns)
    result: dict[tuple[int, str], str | None] = {}

    # Dados começam na linha 2 (linha 1 é cabeçalho)
    for row_i in range(len(df)):
        for col_j, col_name in enumerate(col_names):
            cell = ws.cell(row=row_i + 2, column=col_j + 1)
            fg = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else "00000000"
            result[(row_i, col_name)] = _normalise_hex(str(fg))

    return result


# ---------------------------------------------------------------------------
# Testes de Paridade
# ---------------------------------------------------------------------------


def test_parity_grupo_row(parity_df: pd.DataFrame) -> None:
    """A linha de Grupo deve ter fundo #000000 em ambos os renderers."""
    web_colors = _extract_web_bg_colors(parity_df)
    excel_colors = _extract_excel_bg_colors(parity_df)

    grupo_row_idx = 0  # Primeira linha é o Grupo
    expected_bg = "000000"

    for col in parity_df.columns:
        web_bg = web_colors.get((grupo_row_idx, col))
        excel_bg = excel_colors.get((grupo_row_idx, col))

        assert web_bg == expected_bg, (
            f"Web: linha Grupo, col '{col}': esperado {expected_bg!r}, obtido {web_bg!r}"
        )
        assert excel_bg == expected_bg, (
            f"Excel: linha Grupo, col '{col}': esperado {expected_bg!r}, obtido {excel_bg!r}"
        )


def test_parity_nao_comprar(parity_df: pd.DataFrame) -> None:
    """Linha Não Comprar: colunas CÓDIGO até T Uni devem ter fundo #E6D5F5."""
    web_colors = _extract_web_bg_colors(parity_df)
    excel_colors = _extract_excel_bg_colors(parity_df)

    nao_comprar_row_idx = 1
    expected_bg = _normalise_hex("#E6D5F5")

    # Determinar colunas-alvo da regra Não Comprar
    from orders_master.formatting.rules import RULES
    nao_comprar_rule = next(r for r in RULES if r.name == "Não Comprar")
    target_cols = nao_comprar_rule.target_cells(parity_df)

    assert target_cols, "Regra 'Não Comprar' não retornou colunas-alvo"

    for col in target_cols:
        web_bg = web_colors.get((nao_comprar_row_idx, col))
        excel_bg = excel_colors.get((nao_comprar_row_idx, col))

        assert web_bg == expected_bg, (
            f"Web: linha Não Comprar, col '{col}': esperado {expected_bg!r}, obtido {web_bg!r}"
        )
        assert excel_bg == expected_bg, (
            f"Excel: linha Não Comprar, col '{col}': esperado {expected_bg!r}, obtido {excel_bg!r}"
        )


def test_parity_rutura(parity_df: pd.DataFrame) -> None:
    """Linha Rutura: coluna Proposta deve ter fundo #FF0000."""
    web_colors = _extract_web_bg_colors(parity_df)
    excel_colors = _extract_excel_bg_colors(parity_df)

    rutura_row_idx = 2
    expected_bg = _normalise_hex("#FF0000")
    col = Columns.PROPOSTA

    web_bg = web_colors.get((rutura_row_idx, col))
    excel_bg = excel_colors.get((rutura_row_idx, col))

    assert web_bg == expected_bg, (
        f"Web: linha Rutura, col '{col}': esperado {expected_bg!r}, obtido {web_bg!r}"
    )
    assert excel_bg == expected_bg, (
        f"Excel: linha Rutura, col '{col}': esperado {expected_bg!r}, obtido {excel_bg!r}"
    )


def test_parity_validade_curta(parity_df: pd.DataFrame) -> None:
    """Linha Validade Curta: coluna DTVAL deve ter fundo #FFA500."""
    web_colors = _extract_web_bg_colors(parity_df)
    excel_colors = _extract_excel_bg_colors(parity_df)

    validade_row_idx = 3
    expected_bg = _normalise_hex("#FFA500")
    col = Columns.DTVAL

    web_bg = web_colors.get((validade_row_idx, col))
    excel_bg = excel_colors.get((validade_row_idx, col))

    assert web_bg == expected_bg, (
        f"Web: linha Validade Curta, col '{col}': esperado {expected_bg!r}, obtido {web_bg!r}"
    )
    assert excel_bg == expected_bg, (
        f"Excel: linha Validade Curta, col '{col}': esperado {expected_bg!r}, obtido {excel_bg!r}"
    )


def test_parity_full_matrix(parity_df: pd.DataFrame) -> None:
    """
    Teste mestre: compara TODAS as células entre web e Excel.
    Se web e Excel divergirem em qualquer célula, falha com mensagem clara.
    Este teste é o guardião do SSOT — qualquer nova regra adicionada a rules.py
    que não seja implementada em ambos os renderers irá falhar aqui.
    """
    web_colors = _extract_web_bg_colors(parity_df)
    excel_colors = _extract_excel_bg_colors(parity_df)

    divergences: list[str] = []

    for (row_i, col_name), web_bg in web_colors.items():
        excel_bg = excel_colors.get((row_i, col_name))

        if web_bg != excel_bg:
            row_label = parity_df.at[row_i, Columns.DESIGNACAO] if Columns.DESIGNACAO in parity_df.columns else str(row_i)
            divergences.append(
                f"  Linha {row_i} ('{row_label}'), col '{col_name}': web={web_bg!r} ≠ excel={excel_bg!r}"
            )

    assert not divergences, (
        f"Divergências de cor entre Web Styler e Excel Formatter "
        f"({len(divergences)} célula(s)):\n" + "\n".join(divergences)
    )


def test_normal_row_has_no_highlight(parity_df: pd.DataFrame) -> None:
    """A linha Normal (sem nenhuma regra activa) não deve ter cor de fundo."""
    web_colors = _extract_web_bg_colors(parity_df)
    excel_colors = _extract_excel_bg_colors(parity_df)

    normal_row_idx = 5  # Última linha

    for col in parity_df.columns:
        web_bg = web_colors.get((normal_row_idx, col))
        excel_bg = excel_colors.get((normal_row_idx, col))

        assert web_bg is None, (
            f"Web: linha Normal, col '{col}' não devia ter cor mas tem {web_bg!r}"
        )
        assert excel_bg is None, (
            f"Excel: linha Normal, col '{col}' não devia ter cor mas tem {excel_bg!r}"
        )
