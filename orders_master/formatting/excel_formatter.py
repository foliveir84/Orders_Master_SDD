"""
Excel Formatter — TASK-42.

Implementa a geração de ficheiros Excel com formatação condicional idêntica
à vista web, utilizando as regras centralizadas em ``rules.py``.
"""

import io
import re
from datetime import datetime
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from orders_master.constants import Columns, GroupLabels
from orders_master.formatting.rules import RULES


def sanitize_filename(s: str) -> str:
    """Remove caracteres especiais de uma string para uso em nomes de ficheiro."""
    return re.sub(r"[^A-Za-z0-9_-]", "", s)


def compute_scope_tag(labs: list[str], codes_file: Any, codes_count: int) -> str:
    """
    Determina a tag descritiva do âmbito do processamento para o nome do ficheiro.

    Regras:
    1. Se houver ficheiro de códigos: "TXT-{n}"
    2. Se houver 1 laboratório: nome do laboratório
    3. Se houver múltiplos laboratórios: "Labs-{n}"
    4. Caso contrário: "GRUPO"
    """
    if codes_file is not None:
        return f"TXT-{codes_count}"
    if len(labs) == 1:
        return sanitize_filename(labs[0])
    if len(labs) > 1:
        return f"Labs-{len(labs)}"
    return "GRUPO"


def apply_excel_rules(ws: Any, df: pd.DataFrame) -> None:
    """
    Aplica as HighlightRules do rules.py a uma worksheet openpyxl.
    """
    # Mapeamento de nome de coluna para índice de coluna (1-based para openpyxl)
    col_map = {col: i + 1 for i, col in enumerate(df.columns)}

    # Iterar sobre as linhas do DataFrame (dados começam na linha 2 do Excel)
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        is_grupo = row.get(Columns.LOCALIZACAO) == GroupLabels.GROUP_ROW

        for rule in RULES:
            # Precedência: linha Grupo só recebe regra 1
            if is_grupo and rule.precedence > 1:
                continue

            if rule.predicate(row):
                target_cols = rule.target_cells(df)
                for col_name in target_cols:
                    if col_name in col_map:
                        cell = ws.cell(row=row_idx, column=col_map[col_name])

                        if rule.excel_fill:
                            cell.fill = rule.excel_fill
                        if rule.excel_font:
                            cell.font = rule.excel_font


def build_excel(df: pd.DataFrame, scope_tag: str) -> tuple[bytes, str]:
    """
    Gera um ficheiro Excel em memória com formatação condicional.

    Args:
        df: DataFrame final para exportar.
        scope_tag: Tag de identificação do âmbito (ex: "Mylan", "Labs-3").

    Returns:
        tuple[bytes, str]: Conteúdo do ficheiro em bytes e o nome do ficheiro sugerido.
    """
    # 1. Gerar nome do ficheiro
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"Sell_Out_{scope_tag}_{timestamp}.xlsx"

    # 2. Criar ficheiro Excel base via Pandas
    # Remover colunas técnicas/auxiliares antes de exportar
    hide_cols = [
        "DIR", "DPR", "DATA_OBS", "TimeDelta", 
        "price_anomaly", "_sort_key", "CLA", "CÓDIGO_STR"
    ]
    df_export = df.drop(columns=[c for c in hide_cols if c in df.columns])

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Propostas")

    # 3. Reabrir com openpyxl para aplicar estilos
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    apply_excel_rules(ws, df)

    # Ajustar largura das colunas (básico)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except (ValueError, TypeError, AttributeError):
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    # 4. Salvar final para bytes
    final_output = io.BytesIO()
    wb.save(final_output)

    return final_output.getvalue(), filename
