from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill

from orders_master.constants import Columns, GroupLabels, Highlight


def months_until_expiry(dtval_str: str) -> float:
    """Helper for parsing MM/YYYY and calculating months until expiry."""
    if pd.isna(dtval_str) or not isinstance(dtval_str, str):
        return 999.0
    try:
        parts = dtval_str.split("/")
        if len(parts) == 2:
            month = int(parts[0])
            year = int(parts[1])
            today = datetime.now()
            diff = (year - today.year) * 12 + (month - today.month)
            return float(diff)
    except ValueError:
        pass
    return 999.0


@dataclass
class HighlightRule:
    name: str
    predicate: Callable[[pd.Series], bool]
    target_cells: Callable[[pd.DataFrame], list[str]]
    css_web: str
    excel_fill: PatternFill | None
    excel_font: Font | None
    precedence: int


def _target_nao_comprar(df: pd.DataFrame) -> list[str]:
    # De CÓDIGO até T Uni
    try:
        start_idx_raw = df.columns.get_loc(Columns.CODIGO)
        end_idx_raw = df.columns.get_loc(Columns.T_UNI)
        if not isinstance(start_idx_raw, int) or not isinstance(end_idx_raw, int):
            return []
        start_idx = int(start_idx_raw)
        end_idx = int(end_idx_raw)
        return [str(c) for c in df.columns[start_idx : end_idx + 1]]
    except KeyError:
        return []


RULES: list[HighlightRule] = [
    HighlightRule(
        name="Grupo",
        predicate=lambda s: s.get(Columns.LOCALIZACAO) == GroupLabels.GROUP_ROW,
        target_cells=lambda df: list(df.columns),
        css_web=f"background-color: {Highlight.GRUPO_BG}; color: {Highlight.GRUPO_FG}; font-weight: bold",
        excel_fill=PatternFill(
            start_color=Highlight.GRUPO_BG.replace("#", ""),
            end_color=Highlight.GRUPO_BG.replace("#", ""),
            fill_type="solid",
        ),
        excel_font=Font(color=Highlight.GRUPO_FG.replace("#", ""), bold=True),
        precedence=1,
    ),
    HighlightRule(
        name="Não Comprar",
        predicate=lambda s: pd.notna(s.get(Columns.DATA_OBS)),
        target_cells=_target_nao_comprar,
        css_web=f"background-color: {Highlight.NAO_COMPRAR_BG};",
        excel_fill=PatternFill(
            start_color=Highlight.NAO_COMPRAR_BG.replace("#", ""),
            end_color=Highlight.NAO_COMPRAR_BG.replace("#", ""),
            fill_type="solid",
        ),
        excel_font=None,
        precedence=2,
    ),
    HighlightRule(
        name="Rutura",
        predicate=lambda s: pd.notna(s.get(Columns.DIR)),
        target_cells=lambda df: [Columns.PROPOSTA] if Columns.PROPOSTA in df.columns else [],
        css_web=f"background-color: {Highlight.RUTURA_BG}; color: {Highlight.RUTURA_FG}; font-weight: bold",
        excel_fill=PatternFill(
            start_color=Highlight.RUTURA_BG.replace("#", ""),
            end_color=Highlight.RUTURA_BG.replace("#", ""),
            fill_type="solid",
        ),
        excel_font=Font(color=Highlight.RUTURA_FG.replace("#", ""), bold=True),
        precedence=3,
    ),
    HighlightRule(
        name="Validade Curta",
        predicate=lambda s: months_until_expiry(str(s.get(Columns.DTVAL))) <= 4,
        target_cells=lambda df: [Columns.DTVAL] if Columns.DTVAL in df.columns else [],
        css_web=f"background-color: {Highlight.VALIDADE_BG}; font-weight: bold",
        excel_fill=PatternFill(
            start_color=Highlight.VALIDADE_BG.replace("#", ""),
            end_color=Highlight.VALIDADE_BG.replace("#", ""),
            fill_type="solid",
        ),
        excel_font=Font(bold=True),
        precedence=4,
    ),
    HighlightRule(
        name="Preço Anómalo",
        predicate=lambda s: bool(s.get(Columns.PRICE_ANOMALY, False)),
        target_cells=lambda df: (
            [Columns.PVP_MEDIO]
            if Columns.PVP_MEDIO in df.columns
            else ([Columns.PVP] if Columns.PVP in df.columns else [])
        ),
        css_web="color: #FF0000; font-weight: bold",
        excel_fill=None,
        excel_font=Font(color="FF0000", bold=True),
        precedence=5,
    ),
]

RULES.sort(key=lambda r: r.precedence)
